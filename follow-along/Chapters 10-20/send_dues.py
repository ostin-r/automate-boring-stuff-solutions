'''
send_dues.py is a follow-along project from Chapter 18 of ATBS.
It reads an xlsx file containing information on if a member of
a volunteer group has paid their dues for the month.  If they 
haven't, it will send them an email reminder.
'''
import openpyxl, sys, os
import ezgmail as gmail
import logging as log

log.basicConfig(level=log.DEBUG, format='%(asctime)s: %(message)s')
log.disable(log.CRITICAL)

# open the spreadsheet and get the latest dues status
os.chdir('follow-along/Chapters 10-20')
wb = openpyxl.load_workbook('duesRecords.xlsx')
sheet = wb['Sheet1']
max_col = sheet.max_column
latest_month = sheet.cell(row=1, column=max_col).value.strftime('%B')

# add any members to unpaid list if recent month does not contain 'paid'
unpaid_members = {}
for r in range(2, sheet.max_row+1):
    payment = sheet.cell(row=r, column=max_col).value
    if payment != 'paid':
        member = sheet.cell(row=r, column=1).value
        email = sheet.cell(row=r, column=2).value
        unpaid_members[member] = email

log.debug(f'unpaid members{unpaid_members}')

# send email to members who haven't paid
gmail.init()
for member, email in unpaid_members.items():
    print(f'sending email to {member}...')
    subject = f'{latest_month} cool kid dues unpaid'
    body = f'Dear {member},\n\nThis is a reminder that you have not paid dues for {latest_month}. Please pay them at your earliest convenience \nBest,\nAustin \n\nThis message is sent with python, please report any problems by replying directly.'
    gmail.send(email, subject, body)

print('Done')