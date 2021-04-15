'''
Austin Richards 4/14/21

excel_to_csv.py automates the conversion of many xlsx
files into csv files.  This program names the csv file
in the format <filename>_<sheetname>.csv
'''
import logging
import os, csv, openpyxl
from pathlib import Path
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.DEBUG, format='%(asctime)s: %(message)s')

def get_all_paths(directory):
    '''
    returns all paths in a directory (and it's sub-directories)
    in a list type
    '''
    file_paths = []
    for path, dirs, files in os.walk(directory):
        for filename in files:
            filepath = os.path.join(path, filename)
            file_paths.append(filepath)
    return file_paths


def excel_to_csv(directory):
    # get the absolute path, make a folder within it to save converted files, get files to convert
    directory = os.path.abspath(directory)
    new_dir = os.path.join(directory, 'converted_csv_files')
    os.makedirs(new_dir, exist_ok=True)
    all_paths = [path for path in get_all_paths(directory) if path.endswith('.xlsx')]
    
    for excel_file in all_paths:

        workbook = openpyxl.load_workbook(excel_file)
        excel_filename = Path(excel_file).stem
        print(f'copying {excel_filename}...')

        for sheet_name in workbook.sheetnames:
            print(f'    copying {sheet_name}...')

            # create a csv filename with the excel filename and sheetname, put in new folder
            csv_filename = f'{excel_filename}_{sheet_name}.csv'
            csv_filepath = os.path.join(new_dir, csv_filename)
            new_file = open(csv_filepath, 'w', newline='')
            csv_writer = csv.writer(new_file)

            # get data from the xlsx file and write it to the new csv
            sheet = workbook[sheet_name]
            for row_num in range(1, sheet.max_row + 1):
                row_data = []
                for col_num in range(1, sheet.max_column + 1):
                    col_letter = get_column_letter(col_num)
                    row_data.append(sheet[col_letter + str(row_num)].value)

                # write data to the new csv
                csv_writer.writerow(row_data)
            new_file.close()
    print('files copied.')


excel_to_csv('Chapter 16')