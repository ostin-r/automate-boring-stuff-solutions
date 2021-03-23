'''
blank_row_inserter.py takes an excel file and
inserts 'm' blank rows at row 'n' (inbetween n and n+1)
'''
import os, re
import openpyxl as xl
from openpyxl.utils import get_column_letter
import logging as log

log.basicConfig(level=log.DEBUG, format='%(asctime)s: %(message)s')
log.disable(log.CRITICAL)


def insert_blanks(file, n, m):

    wb = xl.load_workbook(file)
    sheet = wb.active

    # copy all data down +m rows
    print('copying data...')
    for row in range(sheet.max_row, 1, -1): # decrement from bottom row (no data loss)
       if row > n:
            for col in range(1, sheet.max_column + 1):
                letter = get_column_letter(col)
                loc = letter + str(row)
                val = sheet[loc].value

                if str(val).startswith('='):
                    val = move_excel_function(val, m)
                
                new_loc = letter + str(row + m)
                sheet[new_loc].value = val

                log.debug(f'Copying {val} from {loc} to {new_loc}')
    
    print('inserting blank rows...')
    for row in range(n + 1, n + m + 1):
        for col in range(1, sheet.max_column + 1):
            letter = get_column_letter(col)
            loc = letter + str(row)
            sheet[loc].value = ''
            log.debug(f'blank value at {loc}')

    print('saved new file')
    wb.save('insertBlanks.xlsx')


def move_excel_function(value, m):
    '''
    moves a function excel and maintains relative reference for rows
    '''
    loc_regex = re.compile(r'([A-Z]+)(\$?)(\d+)')
    locs = loc_regex.findall(value)
    
    for loc in locs:
        if loc[1]: pass
        old_loc = loc[0] + loc[2]
        new_row = str(int(loc[2]) + m)
        new_loc = loc[0] + new_row
        value = value.replace(old_loc, new_loc, 1)

    return value


os.chdir('Chapter 13')
insert_blanks('.xlsx', 1, 3)