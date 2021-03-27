'''
Austin Richards 3/27/21

txt_xlsx.py contains two functions:
txt_to_xlsx() which takes the contents of multiple .txt files and puts them into an excel file
xlsx_to_txt() which does the reverse of txt_to_xlsx()
'''
import os
import logging as log
import openpyxl as xl
from pathlib import Path as path

log.basicConfig(level=log.DEBUG, format='%(asctime)s : %(message)s')


def txt_to_xlsx(file_list, new_name):
    wb = xl.Workbook()
    sheet = wb.active
    count = 0

    for file in file_list:
        file = open(path(file))
        contents = file.readlines()
        begin = sheet.max_row # begin putting .txt data at this row
        end = sheet.max_row + len(contents) # end putting text data here

        if count > 0:
            begin += 1
            end += 1

        for row in range(begin, end):
            sheet['A' + str(row)].value = contents[row - begin]
            log.debug(f'inserting data {contents[row-begin]} at: {row}')

        count += 1

    wb.save(new_name + '.xlsx')


def xlsx_to_txt(file, new_name):
    wb = xl.load_workbook(file)
    sheet = wb.active
    new_name += '.txt'

    with open(path(new_name), 'w') as file:
        for row in range(1, sheet.max_row + 1):
            data = sheet['A' + str(row)].value
            if not data.endswith('\n'): data += '\n' # organize all lines with newline
            file.write(data)


os.chdir('Chapter 13')
file_list = ['file-1.txt', 'file-2.txt']
txt_to_xlsx(file_list, 'files1&2')
xlsx_to_txt('files1&2.xlsx', 'files1&2')