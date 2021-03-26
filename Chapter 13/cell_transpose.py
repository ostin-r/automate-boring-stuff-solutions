'''
Austin Richards 3/26/21

cell_transpose.py transposes all columns in a sheet
'''
import os
import logging as log
import openpyxl as xl
from openpyxl.utils import get_column_letter

log.basicConfig(level=log.DEBUG, format='%(asctime)s : %(message)s')


def transpose_all(file):
    wb = xl.load_workbook(file)
    sheet = wb.active
    data = [[] for x in range(sheet.max_column)] # a list of lists to hold column data

    new_rows = sheet.max_column # clear up the confusion of transposing...
    new_cols = sheet.max_row    # also important to get these numbers before cutting/pasting data

    # 'cut' data into a list
    for col in range(1, sheet.max_column + 1):
        col_letter = get_column_letter(col)

        for row in range(1, sheet.max_row + 1):
            current_value = sheet[col_letter + str(row)].value
            sheet[col_letter + str(row)].value = '' # cut
            data[col - 1].append(current_value)     # store in list

    # 'paste' in new locations
    for row in range(1, new_rows + 1):
        for col in range(1, new_cols + 1):
            col_letter = get_column_letter(col)
            new_value = data[row - 1][col - 1] # note that rows/cols are flipped 
            sheet[col_letter + str(row)].value = new_value

    print('Saving file...')
    wb.save('transpose-example.xlsx')
    print('Done!')


os.chdir('Chapter 13')
transpose_all('transpose-this.xlsx')