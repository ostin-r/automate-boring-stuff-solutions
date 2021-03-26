'''
Austin Richards 3/24/21

multiply_table_maker.py creates an n x n multiplication table in
an excel file named 'multiplicationTable-n.xlsx'
'''
import os
import openpyxl as xl
import logging as log
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

log.basicConfig(level=log.DEBUG, format='%(asctime)s: %(message)s')


def excel_table(n):
    wb = xl.Workbook()
    sheet = wb.active
    bold_font = Font(bold=True)

    for i in range(1, n + 1):
        letter = get_column_letter(i + 1)
        number = str(i + 1)
        sheet[letter + '1'].value = i
        sheet['A' + number].value = i

        sheet[letter + '1'].font = bold_font
        sheet['A' + number].font = bold_font

    for i in range(1, sheet.max_column):
        for j in range(1, sheet.max_row):
            letter = get_column_letter(i + 1)
            number = str(j + 1)
            sheet[letter + number].value = i * j

    print('table complete!')
    wb.save(f'multiplicationTable-{n}.xlsx')


os.chdir('Chapter 13')
excel_table(10)